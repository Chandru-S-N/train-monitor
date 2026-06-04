import datetime
from io import BytesIO
from django.utils import timezone
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apps.trains.models import Train
from apps.sensors.models import SensorData
from apps.alerts.models import Alert

def generate_pdf_report(report_type, train_id=None, days=1):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#0d0e21'),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#6b7280'),
        spaceAfter=20
    )
    heading_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#060714'),
        spaceBefore=15,
        spaceAfter=10
    )
    body_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1f2937')
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.white
    )

    # Date calculations
    end_date = timezone.now()
    start_date = end_date - datetime.timedelta(days=days)
    
    # Data Querying
    sensors_qs = SensorData.objects.filter(timestamp__range=(start_date, end_date))
    alerts_qs = Alert.objects.filter(timestamp__range=(start_date, end_date))
    
    train_name = "All Fleet"
    if train_id:
        train = Train.objects.filter(id=train_id).first()
        if train:
            train_name = train.name
            sensors_qs = sensors_qs.filter(train=train)
            alerts_qs = alerts_qs.filter(train=train)

    # Page Header
    story.append(Paragraph(f"Smart Train Monitoring Report - {report_type.upper()}", title_style))
    story.append(Paragraph(f"Generated on: {end_date.strftime('%d-%b-%Y %H:%M:%S UTC')} | Target Train: {train_name} | Period: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}", subtitle_style))
    story.append(Spacer(1, 10))

    # Aggregates / Stats Block
    total_readings = sensors_qs.count()
    total_alerts = alerts_qs.count()
    unresolved_alerts = alerts_qs.filter(is_resolved=False).count()
    
    import django.db.models as db_models
    stats = sensors_qs.aggregate(
        avg_temp=db_models.Avg('temperature'),
        avg_pres=db_models.Avg('pressure'),
        avg_vib=db_models.Avg('vibration'),
        avg_speed=db_models.Avg('speed')
    )
    
    avg_temp = stats['avg_temp'] or 0.0
    avg_pres = stats['avg_pres'] or 0.0
    avg_vib = stats['avg_vib'] or 0.0
    avg_speed = stats['avg_speed'] or 0.0

    stats_data = [
        [
            Paragraph("<b>Metric</b>", header_style), 
            Paragraph("<b>Value</b>", header_style), 
            Paragraph("<b>Metric</b>", header_style), 
            Paragraph("<b>Value</b>", header_style)
        ],
        [
            Paragraph("Total Logs Ingested", body_style), 
            Paragraph(str(total_readings), body_style),
            Paragraph("Total Alerts Triggered", body_style), 
            Paragraph(str(total_alerts), body_style)
        ],
        [
            Paragraph("Average Temperature", body_style), 
            Paragraph(f"{avg_temp:.1f} °C", body_style),
            Paragraph("Unresolved Alerts", body_style), 
            Paragraph(str(unresolved_alerts), body_style)
        ],
        [
            Paragraph("Average Pressure", body_style), 
            Paragraph(f"{avg_pres:.1f} psi", body_style),
            Paragraph("Average Speed", body_style), 
            Paragraph(f"{avg_speed:.1f} km/h", body_style)
        ],
        [
            Paragraph("Average Vibration", body_style), 
            Paragraph(f"{avg_vib:.2f} g", body_style),
            Paragraph("", body_style), 
            Paragraph("", body_style)
        ]
    ]
    
    stats_table = Table(stats_data, colWidths=[150, 120, 150, 120])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0d0e21')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f9fafb')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        ('TOPPADDING', (0,1), (-1,-1), 5),
    ]))
    
    story.append(Paragraph("Key Summary Statistics", heading_style))
    story.append(stats_table)
    story.append(Spacer(1, 20))

    # Alerts Breakdown Table (Up to 15)
    story.append(Paragraph("Recent Alert Activities (Max 15)", heading_style))
    alerts_data = [[
        Paragraph("<b>Train</b>", header_style),
        Paragraph("<b>Alert Type</b>", header_style),
        Paragraph("<b>Severity</b>", header_style),
        Paragraph("<b>Status</b>", header_style),
        Paragraph("<b>Timestamp</b>", header_style)
    ]]
    
    for alert in alerts_qs.select_related('train')[:15]:
        status_str = "Resolved" if alert.is_resolved else ("Acked" if alert.is_acknowledged else "New")
        alerts_data.append([
            Paragraph(alert.train.name, body_style),
            Paragraph(alert.get_alert_type_display(), body_style),
            Paragraph(alert.severity.upper(), body_style),
            Paragraph(status_str, body_style),
            Paragraph(alert.timestamp.strftime('%d-%b %H:%M'), body_style),
        ])
        
    if len(alerts_data) == 1:
        alerts_data.append([Paragraph("No alerts recorded during this timeframe.", body_style), "", "", "", ""])

    alerts_table = Table(alerts_data, colWidths=[110, 150, 80, 80, 120])
    alerts_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3b82f6')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f9fafb')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(alerts_table)
    story.append(Spacer(1, 20))

    # Historical Logs Preview (Up to 25)
    story.append(Paragraph("Recent Ingestion History (Max 25)", heading_style))
    readings_data = [[
        Paragraph("<b>Train</b>", header_style),
        Paragraph("<b>Temp (°C)</b>", header_style),
        Paragraph("<b>Pressure (psi)</b>", header_style),
        Paragraph("<b>Vibration (g)</b>", header_style),
        Paragraph("<b>Smoke</b>", header_style),
        Paragraph("<b>Speed (km/h)</b>", header_style),
        Paragraph("<b>Timestamp</b>", header_style)
    ]]
    
    for row in sensors_qs.select_related('train')[:25]:
        readings_data.append([
            Paragraph(row.train.id, body_style),
            Paragraph(f"{row.temperature:.1f}", body_style),
            Paragraph(f"{row.pressure:.1f}", body_style),
            Paragraph(f"{row.vibration:.3f}", body_style),
            Paragraph(f"{row.smoke:.1f}", body_style),
            Paragraph(f"{row.speed:.1f}", body_style),
            Paragraph(row.timestamp.strftime('%d-%b %H:%M:%S'), body_style),
        ])
        
    if len(readings_data) == 1:
        readings_data.append([Paragraph("No telemetry records logged during this timeframe.", body_style), "", "", "", "", "", ""])

    readings_table = Table(readings_data, colWidths=[70, 70, 80, 80, 60, 80, 100])
    readings_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#8b5cf6')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f9fafb')),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(readings_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(report_type, train_id=None, days=1):
    wb = Workbook()
    
    # Styles
    title_font = Font(name='Segoe UI', size=16, bold=True, color='0D0E21')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    section_font = Font(name='Segoe UI', size=13, bold=True, color='1F2937')
    bold_font = Font(name='Segoe UI', size=10, bold=True)
    regular_font = Font(name='Segoe UI', size=10)
    
    header_fill = PatternFill(start_color='0D0E21', end_color='0D0E21', fill_type='solid')
    stats_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Date calculations
    end_date = timezone.now()
    start_date = end_date - datetime.timedelta(days=days)
    
    # Data Querying
    sensors_qs = SensorData.objects.filter(timestamp__range=(start_date, end_date)).select_related('train')
    alerts_qs = Alert.objects.filter(timestamp__range=(start_date, end_date)).select_related('train')
    
    train_name = "All Fleet"
    if train_id:
        train = Train.objects.filter(id=train_id).first()
        if train:
            train_name = train.name
            sensors_qs = sensors_qs.filter(train=train)
            alerts_qs = alerts_qs.filter(train=train)

    # 1. Sheet 1: Summary Statistics
    ws_summary = wb.active
    ws_summary.title = "Summary & KPIs"
    ws_summary.views.sheetView[0].showGridLines = True
    
    ws_summary['A1'] = "Smart Train Monitoring Report"
    ws_summary['A1'].font = title_font
    ws_summary['A2'] = f"Generated: {end_date.strftime('%d-%b-%Y %H:%M:%S UTC')} | Train: {train_name} | Range: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
    ws_summary['A2'].font = Font(name='Segoe UI', size=10, italic=True, color='6B7280')
    
    ws_summary['A4'] = "Overview Metrics"
    ws_summary['A4'].font = section_font
    
    # Calculate stats
    total_readings = sensors_qs.count()
    total_alerts = alerts_qs.count()
    unresolved_alerts = alerts_qs.filter(is_resolved=False).count()
    
    import django.db.models as db_models
    stats = sensors_qs.aggregate(
        avg_temp=db_models.Avg('temperature'),
        avg_pres=db_models.Avg('pressure'),
        avg_vib=db_models.Avg('vibration'),
        avg_speed=db_models.Avg('speed')
    )
    
    avg_temp = stats['avg_temp'] or 0.0
    avg_pres = stats['avg_pres'] or 0.0
    avg_vib = stats['avg_vib'] or 0.0
    avg_speed = stats['avg_speed'] or 0.0

    kpis = [
        ("Total Telemetry Logs Ingested", total_readings, ""),
        ("Total Warnings/Alerts Logged", total_alerts, "alerts"),
        ("Active Unresolved Alerts", unresolved_alerts, "active"),
        ("Fleet Average Temperature", round(avg_temp, 2), "°C"),
        ("Fleet Average Pressure", round(avg_pres, 2), "psi"),
        ("Fleet Average Speed", round(avg_speed, 2), "km/h"),
        ("Fleet Average Vibration", round(avg_vib, 3), "g"),
    ]
    
    row = 5
    for label, val, unit in kpis:
        ws_summary.cell(row=row, column=1, value=label).font = regular_font
        ws_summary.cell(row=row, column=2, value=f"{val} {unit}".strip()).font = bold_font
        ws_summary.cell(row=row, column=1).fill = stats_fill
        ws_summary.cell(row=row, column=2).fill = stats_fill
        ws_summary.cell(row=row, column=1).border = thin_border
        ws_summary.cell(row=row, column=2).border = thin_border
        row += 1

    # 2. Sheet 2: Telemetry Data Logs
    ws_logs = wb.create_sheet(title="Telemetry Logs")
    ws_logs.views.sheetView[0].showGridLines = True
    
    log_headers = ["Train ID", "Train Name", "Temp (°C)", "Pressure (psi)", "Vibration (g)", "Smoke/Gas", "Latitude", "Longitude", "Speed (km/h)", "Status", "Timestamp"]
    for col_idx, text in enumerate(log_headers, 1):
        cell = ws_logs.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(sensors_qs[:5000], 2):
        ws_logs.cell(row=row_idx, column=1, value=item.train.id).font = regular_font
        ws_logs.cell(row=row_idx, column=2, value=item.train.name).font = regular_font
        ws_logs.cell(row=row_idx, column=3, value=item.temperature).font = regular_font
        ws_logs.cell(row=row_idx, column=4, value=item.pressure).font = regular_font
        ws_logs.cell(row=row_idx, column=5, value=item.vibration).font = regular_font
        ws_logs.cell(row=row_idx, column=6, value=item.smoke).font = regular_font
        ws_logs.cell(row=row_idx, column=7, value=item.latitude).font = regular_font
        ws_logs.cell(row=row_idx, column=8, value=item.longitude).font = regular_font
        ws_logs.cell(row=row_idx, column=9, value=item.speed).font = regular_font
        ws_logs.cell(row=row_idx, column=10, value=item.status).font = regular_font
        ws_logs.cell(row=row_idx, column=11, value=item.timestamp.replace(tzinfo=None)).font = regular_font
        
        # Apply borders to details
        for c in range(1, 12):
            ws_logs.cell(row=row_idx, column=c).border = thin_border

    # Auto-adjust column widths
    for ws in [ws_summary, ws_logs]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
